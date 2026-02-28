# -*- coding: utf-8 -*-
"""
employee_bulk_upload_wizard.py
-------------------------------
Wizard to bulk-import employees from an Excel (.xlsx) file.

Design notes
------------
* Depends on beto_hr which already defines:
    employee_id_no, device_user_id, joining_date, supervisor_id,
    dotted_supervisor_id, religion_id, blood_group_id,
    permanent_address, present_address, personal_email
  on hr.employee.

* Excel columns → Odoo field mapping is declared in COLUMN_MAP.

* Validation runs entirely in-memory (no DB write until user confirms).

* Each confirmed row is written inside a savepoint so that a single bad
  row doesn't abort a partial-import run.

* res.users is created with only base.group_user and no signup email.
"""

import base64
import io
import json
import logging
import random
import string
from datetime import date, datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _random_password(length=12):
    """Generate a secure random password with letters and digits."""
    chars = string.ascii_letters + string.digits
    return ''.join(random.SystemRandom().choice(chars) for _ in range(length))


def _parse_date(raw):
    """
    Accept Excel date-serial (float/int), 'YYYY-MM-DD', or 'DD/MM/YYYY'.
    Returns a date object or None.
    """
    if raw is None or raw == '':
        return None
    if isinstance(raw, (datetime,)):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, float):
        # Excel serial date (openpyxl with data_only=True gives floats for
        # cells formatted as dates when no type hint is available)
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(raw).date()
        except Exception:
            pass
    if isinstance(raw, int):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(raw).date()
        except Exception:
            pass
    raw = str(raw).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None  # unparseable


def _parse_str(raw):
    """
    Safely convert Excel values (like floats from numeric cells) to string,
    removing .0 suffixes for phone numbers/IDs.
    """
    if raw is None or raw == '':
        return ''
    if isinstance(raw, float):
        if raw.is_integer():
            raw = int(raw)
    result = str(raw).strip()
    # Pro fix: Auto-prepend '0' to BD phone numbers if Excel stripped the leading zero (10-digits starting with '1')
    if len(result) == 10 and result.startswith('1'):
        result = '0' + result
    return result


_GENDER_MAP = {
    'male': 'male',
    'female': 'female',
    'other': 'other',
}

# ---------------------------------------------------------------------------
# Column → field map
# Each entry: Excel header → (odoo_field, field_type)
# field_type: 'char', 'date', 'gender', 'm2o_company', 'm2o_department',
#             'm2o_job', 'm2o_employee', 'm2o_religion', 'm2o_blood_group',
#             'selection'
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    'Business Unit':         ('company_id', 'm2o_company'),
    'Employee Name':         ('name', 'char'),
    'Employee ID':           ('employee_id_no', 'char'),
    'Device ID':             ('device_user_id', 'char'),
    'Joining Date':          ('joining_date', 'date'),
    'Department':            ('department_id', 'm2o_department'),
    'Designation':           ('job_id', 'm2o_job'),
    'Supervisor':            ('supervisor_id', 'm2o_employee'),
    'Dotted Supervisor':     ('dotted_supervisor_id', 'm2o_employee'),
    'Line Manager':          ('parent_id', 'm2o_employee'),
    'Work Email':            ('work_email', 'char'),
    'Work Mobile':           ('mobile_phone', 'char'),
    'Work Phone Number':     ('work_phone', 'char'),
    'Personal Phone':        ('private_phone', 'char'),
    'Private Email':         ('private_email', 'char'),
    'Permanent Address':     ('permanent_address', 'char'),
    'Present Address':       ('present_address', 'char'),
    'Date of Birth':         ('birthday', 'date'),
    'Gender':                ('sex', 'gender'),
    'Religion':              ('religion_id', 'm2o_religion'),
    'Blood Group':           ('blood_group_id', 'm2o_blood_group'),
    'NID No':                ('identification_id', 'char'),
    'TIN Number':            ('tin_number', 'char'),
    'Place of Birth':        ('place_of_birth', 'char'),
    'Nationality (Country)': ('country_id', 'char'),
    'Employee Type':         ('employee_type', 'char'),
    'Employment Type':       ('employee_category', 'char'),
    'Bank Accounts':         ('bank_account_ids', 'char'),
}

# Excel header label → Odoo field mapping for required check
REQUIRED_FIELDS = {
    # Core identity
    'company_id', 'name', 'employee_id_no', 'device_user_id', 'work_email',
    # Dates 
    'joining_date',
    # Personal details 
    'permanent_address', 'present_address', 'birthday', 'sex',
    'religion_id', 'identification_id', 'place_of_birth', 'country_id',
    # Employment classification 
    'employee_type', 'employee_category',
}
# Column headers that must be required (derived from COLUMN_MAP + REQUIRED_FIELDS)
REQUIRED_HEADERS = {
    hdr for hdr, (field, _) in COLUMN_MAP.items() if field in REQUIRED_FIELDS
}


# ===========================================================================
# TransientModel: one line per Excel row
# ===========================================================================
class EmployeeBulkUploadLine(models.TransientModel):
    _name = 'employee.bulk.upload.line'
    _description = 'Employee Bulk Upload Preview Line'
    _order = 'row_no'

    wizard_id = fields.Many2one(
        'employee.bulk.upload.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True,
    )
    row_no = fields.Integer(string='Row #', readonly=True)
    employee_id_no = fields.Char(string='Employee ID', readonly=True)
    device_user_id = fields.Char(string='Device ID', readonly=True)
    name = fields.Char(string='Employee Name', readonly=True)
    work_email = fields.Char(string='Work Email', readonly=True)
    company = fields.Char(string='Company', readonly=True)
    department = fields.Char(string='Department', readonly=True)
    job = fields.Char(string='Designation', readonly=True)
    is_valid = fields.Boolean(string='Valid', readonly=True)
    import_status = fields.Selection([
        ('pending', 'Pending'),
        ('success', 'Success'),
        ('failed', 'Failed')
    ], string='Import Status', default='pending', readonly=True)
    error_text = fields.Text(string='Errors', readonly=True)
    raw_json = fields.Text(string='Raw Data (JSON)', readonly=True)

    def action_show_error(self):
        self.ensure_one()
        if not self.error_text:
            return True
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Validation Error (Row %s)') % self.row_no,
                'message': self.error_text,
                'sticky': True,
                'type': 'danger',
            }
        }


# ===========================================================================
# TransientModel: main wizard
# ===========================================================================
class EmployeeBulkUploadWizard(models.TransientModel):
    _name = 'employee.bulk.upload.wizard'
    _description = 'Employee Bulk Upload Wizard'

    # -----------------------------------------------------------------------
    # Fields
    # -----------------------------------------------------------------------
    file = fields.Binary(string='Excel File (.xlsx)', required=False, attachment=False)
    filename = fields.Char(string='File Name')

    create_missing_departments = fields.Boolean(
        string='Create Missing Departments',
        default=False,
        help='If checked, departments not found by name will be created automatically.',
    )
    create_missing_jobs = fields.Boolean(
        string='Create Missing Designations',
        default=False,
        help='If checked, job positions not found by name will be created automatically.',
    )
    stop_on_error = fields.Boolean(
        string='Stop on First Error',
        default=False,
        help='If checked, the whole import is rolled back when any row fails. '
             'If unchecked, valid rows are imported via savepoints and invalid rows are skipped.',
    )

    line_ids = fields.One2many(
        'employee.bulk.upload.line',
        'wizard_id',
        string='Preview Lines',
        readonly=True,
    )

    state = fields.Selection([
        ('draft', 'Upload'),
        ('preview', 'Preview'),
        ('done', 'Done'),
    ], default='draft', string='State')

    # Output / log
    import_log_id = fields.Many2one(
        'employee.bulk.import.log',
        string='Import Log',
        readonly=True,
    )
    output_file = fields.Binary(string='Output Excel', readonly=True)
    output_filename = fields.Char(string='Output Filename', readonly=True)
    error_file = fields.Binary(string='Error Report Excel', readonly=True)
    error_filename = fields.Char(string='Error Filename', readonly=True)
    import_summary = fields.Text(string='Import Summary', readonly=True)

    has_errors = fields.Boolean(
        compute='_compute_has_errors',
        string='Has Errors',
    )

    @api.depends('line_ids.is_valid')
    def _compute_has_errors(self):
        for rec in self:
            rec.has_errors = any(not ln.is_valid for ln in rec.line_ids)

    # -----------------------------------------------------------------------
    # Public actions
    # -----------------------------------------------------------------------

    def action_validate(self):
        """Parse the Excel file and populate line_ids with validation results.
        No DB write to hr.employee occurs here."""
        self.ensure_one()
        if not self.file:
            raise UserError(_('Please upload an Excel file first.'))

        rows = self._parse_excel()
        # Clear old lines
        self.line_ids.unlink()

        validated_lines = []
        for row in rows:
            line_vals = self._validate_row(row)
            validated_lines.append(line_vals)

        self.write({
            'line_ids': [(0, 0, v) for v in validated_lines],
            'state': 'preview',
        })
        # Re-open the wizard to show the preview
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'employee.bulk.upload.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_confirm_import(self):
        """Import valid rows. Behaviour controlled by stop_on_error."""
        self.ensure_one()
        if self.state != 'preview':
            raise UserError(_('Please validate the file before importing.'))

        rows = self._parse_excel()
        # Build a lookup from row_no → validated line record
        line_by_row = {ln.row_no: ln for ln in self.line_ids}

        results = []        # list of dicts for output Excel
        failed_rows = []    # list of (row_no, error) for error report

        env = self.env

        if self.stop_on_error:
            # Single transaction – if anything fails, all fails
            for row_data in rows:
                row_no = row_data['_row_no']
                line = line_by_row.get(row_no)
                if line and not line.is_valid:
                    failed_rows.append({'row_no': row_no, 'error': line.error_text})
                    if line:
                        line.import_status = 'failed'
                    raise UserError(
                        _('Row %(row)s has validation errors – aborting import.\n\n%(err)s',
                          row=row_no, err=line.error_text)
                    )
                result = self._import_row(row_data, env)
                if result.get('error'):
                    if line:
                        line.import_status = 'failed'
                        line.error_text = result['error']
                    raise UserError(
                        _('Row %(row)s failed during import: %(err)s',
                          row=row_no, err=result['error'])
                    )
                if line:
                    line.import_status = 'success'
                results.append(result)
        else:
            # Partial import via savepoints
            for row_data in rows:
                row_no = row_data['_row_no']
                line = line_by_row.get(row_no)
                if line and not line.is_valid:
                    failed_rows.append({'row_no': row_no, 'error': line.error_text})
                    if line:
                        line.import_status = 'failed'
                    continue
                try:
                    with env.cr.savepoint():
                        result = self._import_row(row_data, env)
                        if result.get('error'):
                            raise Exception(result['error'])
                        results.append(result)
                        if line:
                            line.import_status = 'success'
                except Exception as exc:
                    _logger.exception('Row %s failed during import', row_no)
                    failed_rows.append({'row_no': row_no, 'error': str(exc)})
                    if line:
                        line.import_status = 'failed'
                        line.error_text = str(exc)

        output_file_b64, output_fname = self._build_output_excel(results)
        error_file_b64, error_fname = self._build_error_excel(failed_rows)

        summary_parts = [
            f'Total rows: {len(rows)}',
            f'Imported:   {len(results)}',
            f'Failed:     {len(failed_rows)}',
        ]
        summary = '\n'.join(summary_parts)

        # Persist in import log
        log_vals = {
            'name': f'Import {fields.Datetime.now()}',
            'imported_count': len(results),
            'failed_count': len(failed_rows),
            'output_file': output_file_b64,
            'output_filename': output_fname,
            'error_file': error_file_b64,
            'error_filename': error_fname,
            'summary': summary,
        }
        log = env['employee.bulk.import.log'].create(log_vals)

        self.write({
            'state': 'done',
            'import_log_id': log.id,
            'output_file': output_file_b64,
            'output_filename': output_fname,
            'error_file': error_file_b64,
            'error_filename': error_fname,
            'import_summary': summary,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'employee.bulk.upload.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_download_output(self):
        """Download the output Excel (password sheet)."""
        self.ensure_one()
        if not self.output_file:
            raise UserError(_('No output file available yet. Please run the import first.'))
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/employee.bulk.upload.wizard/{self.id}/output_file/{self.output_filename}?download=true',
            'target': 'self',
        }

    def action_download_errors(self):
        """Download the error-report Excel."""
        self.ensure_one()
        if not self.error_file:
            raise UserError(_('No error report available.'))
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/employee.bulk.upload.wizard/{self.id}/error_file/{self.error_filename}?download=true',
            'target': 'self',
        }

    def action_download_template(self):
        """Generate and download a styled Excel template.

        - Required column headers: orange background, white bold text.
        - Optional column headers: light grey background, dark bold text.
        - Date columns: pre-formatted with DD/MM/YYYY number format.
        """
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_('openpyxl is required. Install it with: pip install openpyxl'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employees'

        # ---- Style definitions ----
        required_fill = PatternFill(fill_type='solid', fgColor='E65100')   # Deep orange
        optional_fill = PatternFill(fill_type='solid', fgColor='CFD8DC')   # Blue-grey light
        req_font  = Font(bold=True, color='FFFFFF', size=11)
        opt_font  = Font(bold=True, color='1A237E', size=11)
        center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='B0BEC5')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Date columns for pre-formatting
        date_headers = {hdr for hdr, (_, ftype) in COLUMN_MAP.items() if ftype == 'date'}

        headers = list(COLUMN_MAP.keys())
        ws.row_dimensions[1].height = 30  # taller header row

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)

            # Header cell
            cell = ws.cell(row=1, column=col_idx, value=header)
            is_required = header in REQUIRED_HEADERS
            cell.fill      = required_fill if is_required else optional_fill
            cell.font      = req_font      if is_required else opt_font
            cell.alignment = center
            cell.border    = border

            # Column width
            ws.column_dimensions[col_letter].width = 22

            # Pre-format the data rows (2-1000) as date where applicable
            if header in date_headers:
                date_fmt = 'DD/MM/YYYY'
                for row_idx in range(2, 1001):
                    data_cell = ws.cell(row=row_idx, column=col_idx)
                    data_cell.number_format = date_fmt

        # Freeze the header row
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        template_b64 = base64.b64encode(buf.getvalue()).decode()

        self.write({
            'output_file': template_b64,
            'output_filename': 'employee_upload_template.xlsx',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/employee.bulk.upload.wizard/{self.id}/output_file/employee_upload_template.xlsx?download=true',
            'target': 'self',
        }

    # -----------------------------------------------------------------------
    # Private – Parsing
    # -----------------------------------------------------------------------

    def _parse_excel(self):
        """
        Read the uploaded Excel file.
        Returns a list of dicts: {header_name: cell_value, '_row_no': int}
        """
        try:
            import openpyxl
        except ImportError:
            raise UserError(_('openpyxl is required. Install it with: pip install openpyxl'))

        try:
            content = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception as e:
            raise UserError(_('Could not read Excel file: %s') % str(e))

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            raise UserError(_('The Excel file appears to be empty.'))

        # Normalise headers: strip whitespace
        headers = [str(h).strip() if h is not None else '' for h in raw_headers]

        data = []
        for row_idx, row_vals in enumerate(rows_iter, start=2):  # data starts row 2
            # Skip completely empty rows
            if all(v is None or str(v).strip() == '' for v in row_vals):
                continue
            row_dict = {'_row_no': row_idx}
            for col_idx, val in enumerate(row_vals):
                if col_idx < len(headers):
                    hdr = headers[col_idx]
                    row_dict[hdr] = val
            data.append(row_dict)

        wb.close()
        return data

    # -----------------------------------------------------------------------
    # Private – Validation
    # -----------------------------------------------------------------------

    def _validate_row(self, row_data):
        """
        Validate a single row dict.
        Returns a dict suitable for creating employee.bulk.upload.line.
        Does NOT write to the DB.
        """
        errors = []
        row_no = row_data.get('_row_no', '?')

        # Helper to get Excel value by header label
        def get(label):
            return row_data.get(label)

        # ------------------------------------------------------------------
        # Extract raw values
        # ------------------------------------------------------------------
        company_name = _parse_str(get('Business Unit'))
        emp_name = _parse_str(get('Employee Name'))
        emp_id_no = _parse_str(get('Employee ID'))
        device_id = _parse_str(get('Device ID'))
        work_email = _parse_str(get('Work Email'))
        dept_name = _parse_str(get('Department'))
        job_name = _parse_str(get('Designation'))

        # ------------------------------------------------------------------
        # Required field checks
        # ------------------------------------------------------------------
        if not company_name:
            errors.append('Business Unit (company_id) is required.')
        if not emp_name:
            errors.append('Employee Name is required.')
        if not emp_id_no:
            errors.append('Employee ID is required.')
        if not device_id:
            errors.append('Device ID is required.')
        if not work_email:
            errors.append('Work Email is required.')

        # Additional required fields
        if not _parse_str(get('Permanent Address')):
            errors.append('Permanent Address is required.')
        if not _parse_str(get('Present Address')):
            errors.append('Present Address is required.')
        if not get('Date of Birth'):
            errors.append('Date of Birth is required.')
        if not _parse_str(get('Gender')):
            errors.append('Gender is required.')
        if not _parse_str(get('Religion')):
            errors.append('Religion is required.')
        if not _parse_str(get('NID No')):
            errors.append('NID No is required.')
        if not _parse_str(get('Place of Birth')):
            errors.append('Place of Birth is required.')
        if not _parse_str(get('Nationality (Country)')):
            errors.append('Nationality (Country) is required.')
        if not _parse_str(get('Employee Type')):
            errors.append('Employee Type is required.')
        if not _parse_str(get('Employment Type')):
            errors.append('Employment Type is required.')

        # ------------------------------------------------------------------
        # Company resolution
        # ------------------------------------------------------------------
        company_id = None
        if company_name:
            company = self.env['res.company'].sudo().search(
                [('name', '=', company_name)], limit=1
            )
            if not company:
                errors.append(f'Company "{company_name}" not found.')
            else:
                company_id = company.id

        # ------------------------------------------------------------------
        # Uniqueness checks (in existing DB)
        # ------------------------------------------------------------------
        if emp_id_no:
            existing = self.env['hr.employee'].sudo().search(
                [('employee_id_no', '=', emp_id_no)], limit=1
            )
            if existing:
                errors.append(
                    f'Employee ID "{emp_id_no}" already exists (employee: {existing.name}).'
                )
        if device_id:
            existing = self.env['hr.employee'].sudo().search(
                [('device_user_id', '=', device_id)], limit=1
            )
            if existing:
                errors.append(
                    f'Device ID "{device_id}" already exists (employee: {existing.name}).'
                )
        if work_email:
            existing_user = self.env['res.users'].sudo().search(
                [('login', '=', work_email)], limit=1
            )
            if existing_user:
                errors.append(
                    f'Work Email "{work_email}" is already used as login '
                    f'(user: {existing_user.name}).'
                )

        # ------------------------------------------------------------------
        # Also check uniqueness within this upload batch (across other rows)
        # ------------------------------------------------------------------
        for ln in self.line_ids:
            if emp_id_no and ln.employee_id_no == emp_id_no:
                errors.append(
                    f'Duplicate Employee ID "{emp_id_no}" in this upload batch (row {ln.row_no}).'
                )
            if device_id and ln.device_user_id == device_id:
                errors.append(
                    f'Duplicate Device ID "{device_id}" in this upload batch (row {ln.row_no}).'
                )
            if work_email and ln.work_email == work_email:
                errors.append(
                    f'Duplicate Work Email "{work_email}" in this upload batch (row {ln.row_no}).'
                )

        # ------------------------------------------------------------------
        # Department resolution (only validate; actual creation in import)
        # ------------------------------------------------------------------
        if dept_name and company_id:
            dept = self.env['hr.department'].sudo().search(
                [('name', '=', dept_name), ('company_id', '=', company_id)], limit=1
            )
            if not dept and not self.create_missing_departments:
                errors.append(
                    f'Department "{dept_name}" not found under company "{company_name}". '
                    'Enable "Create Missing Departments" or add the department first.'
                )

        # Job resolution
        if job_name:
            job = self.env['hr.job'].sudo().search(
                [('name', '=', job_name)], limit=1
            )
            if not job and not self.create_missing_jobs:
                errors.append(
                    f'Designation "{job_name}" not found. '
                    'Enable "Create Missing Designations" or add it first.'
                )

        # Supervisor / DottedSupervisor / LineManager validation
        for label, field in [
            ('Supervisor', 'supervisor_id'),
            ('Dotted Supervisor', 'dotted_supervisor_id'),
            ('Line Manager', 'parent_id'),
        ]:
            raw_val = str(get(label) or '').strip()
            if raw_val:
                found = self._resolve_employee_ref(raw_val)
                if not found:
                    errors.append(
                        f'{label} "{raw_val}" could not be resolved. '
                        'Provide either Employee ID or Work Email of an existing employee.'
                    )

        # Date validations – Joining Date is required; Date of Birth optional but must be parseable
        joining_date_raw = get('Joining Date')
        if not joining_date_raw or str(joining_date_raw).strip() == '':
            errors.append('Joining Date is required.')
        else:
            if _parse_date(joining_date_raw) is None:
                errors.append(
                    f'Joining Date "{joining_date_raw}" could not be parsed. '
                    'Use DD/MM/YYYY or YYYY-MM-DD format.'
                )
        dob_raw = get('Date of Birth')
        if dob_raw and str(dob_raw).strip() != '':
            if _parse_date(dob_raw) is None:
                errors.append(
                    f'Date of Birth "{dob_raw}" could not be parsed as a date. '
                    'Use DD/MM/YYYY or YYYY-MM-DD format.'
                )

        is_valid = len(errors) == 0

        return {
            'wizard_id': self.id,
            'row_no': row_no,
            'employee_id_no': emp_id_no,
            'device_user_id': device_id,
            'name': emp_name,
            'work_email': work_email,
            'company': company_name,
            'department': dept_name,
            'job': job_name,
            'is_valid': is_valid,
            'error_text': '\n'.join(errors) if errors else '',
            'raw_json': json.dumps(
                {k: str(v) if v is not None else '' for k, v in row_data.items()},
                ensure_ascii=False,
            ),
        }

    # -----------------------------------------------------------------------
    # Private – Import single row
    # -----------------------------------------------------------------------

    def _import_row(self, row_data, env):
        """
        Create/update one hr.employee + res.users.
        Returns a dict with employee info + generated password (for output Excel).
        """
        def get(label):
            return row_data.get(label)

        company_name = _parse_str(get('Business Unit'))
        emp_name = _parse_str(get('Employee Name'))
        emp_id_no = _parse_str(get('Employee ID'))
        device_id = _parse_str(get('Device ID'))
        work_email = _parse_str(get('Work Email'))
        dept_name = _parse_str(get('Department'))
        job_name = _parse_str(get('Designation'))

        # Resolve company
        company = env['res.company'].sudo().search(
            [('name', '=', company_name)], limit=1
        )
        if not company:
            return {'error': f'Company "{company_name}" not found.'}

        # Resolve / create department
        dept_id = False
        if dept_name:
            dept = env['hr.department'].sudo().search(
                [('name', '=', dept_name), ('company_id', '=', company.id)], limit=1
            )
            if not dept and self.create_missing_departments:
                dept = env['hr.department'].sudo().create({
                    'name': dept_name,
                    'company_id': company.id,
                })
            dept_id = dept.id if dept else False

        # Resolve / create job
        job_id = False
        if job_name:
            job = env['hr.job'].sudo().search(
                [('name', '=', job_name)], limit=1
            )
            if not job and self.create_missing_jobs:
                job = env['hr.job'].sudo().create({'name': job_name})
            job_id = job.id if job else False

        # Resolve many2one employees (supervisor / dotted-supervisor / line-manager)
        supervisor_id = self._resolve_employee_ref(_parse_str(get('Supervisor')))
        dotted_supervisor_id = self._resolve_employee_ref(_parse_str(get('Dotted Supervisor')))
        parent_id = self._resolve_employee_ref(_parse_str(get('Line Manager')))

        # Religion
        religion_id = False
        religion_name = _parse_str(get('Religion'))
        if religion_name:
            religion = env['hr.religion'].sudo().search(
                [('name', '=ilike', religion_name)], limit=1
            )
            if not religion:
                religion = env['hr.religion'].sudo().create({'name': religion_name})
            religion_id = religion.id

        # Blood Group
        blood_group_id = False
        blood_group_name = _parse_str(get('Blood Group'))
        if blood_group_name:
            bg = env['hr.blood.group'].sudo().search(
                [('name', '=ilike', blood_group_name)], limit=1
            )
            if not bg:
                bg = env['hr.blood.group'].sudo().create({'name': blood_group_name})
            blood_group_id = bg.id

        # Nationality (Country)
        country_id = False
        country_name = _parse_str(get('Nationality (Country)'))
        if country_name:
            country = env['res.country'].sudo().search(
                [('name', '=ilike', country_name)], limit=1
            )
            if country:
                country_id = country.id

        # Dates
        joining_date = _parse_date(get('Joining Date'))
        birthday = _parse_date(get('Date of Birth'))

        # Gender
        gender_raw = _parse_str(get('Gender')).lower()
        gender = _GENDER_MAP.get(gender_raw, False)

        # Employee Type (Selection)
        emp_type_raw = _parse_str(get('Employee Type'))
        final_emp_type = False
        if emp_type_raw:
            try:
                selection = env['hr.employee']._fields['employee_type'].selection
                if callable(selection):
                    selection = selection(env['hr.employee'])
                # Match case-insensitively on selection keys or labels
                for key, label in selection:
                    if emp_type_raw.lower() == str(key).lower() or emp_type_raw.lower() == str(label).lower():
                        final_emp_type = key
                        break
            except Exception:
                pass

            if not final_emp_type and emp_type_raw.lower() == 'freelancer':
                final_emp_type = 'freelance'
            elif not final_emp_type:
                final_emp_type = emp_type_raw.lower() # Fallback

        # Contract Type / Employment Type
        contract_type_name = _parse_str(get('Employment Type'))
        contract_type_id = False
        if contract_type_name:
            ct = env['hr.contract.type'].sudo().search([('name', '=ilike', contract_type_name)], limit=1)
            if not ct:
                ct = env['hr.contract.type'].sudo().create({'name': contract_type_name})
            contract_type_id = ct.id

        # Build employee vals
        employee_vals = {
            'name': emp_name,
            'company_id': company.id,
            'employee_id_no': emp_id_no,
            'device_user_id': device_id,
            'work_email': work_email,
            'work_phone': _parse_str(get('Work Phone Number')) or False,
            'mobile_phone': _parse_str(get('Work Mobile')) or False,
            'identification_id': _parse_str(get('NID No')) or False,
            'employee_type': final_emp_type or False,
            'contract_type_id': contract_type_id or False,
            'personal_email': _parse_str(get('Personal Email')) or False,
            'private_email': _parse_str(get('Private Email')) or False,
            'private_phone': _parse_str(get('Personal Phone')) or False,
            'permanent_address': str(get('Permanent Address') or '').strip() or False,
            'present_address': str(get('Present Address') or '').strip() or False,
            'department_id': dept_id,
            'job_id': job_id,
            'supervisor_id': supervisor_id or False,
            'dotted_supervisor_id': dotted_supervisor_id or False,
            'parent_id': parent_id or False,
            'religion_id': religion_id or False,
            'blood_group_id': blood_group_id or False,
            'birthday': birthday or False,
            'sex': gender or False,
            'tin_number': _parse_str(get('TIN Number')) or False,
            'place_of_birth': _parse_str(get('Place of Birth')) or False,
            'country_id': country_id or False,
            'hr_responsible_id': env.user.id,  # Always set to the uploader
        }
        if joining_date:
            employee_vals['joining_date'] = joining_date

        # Check employee_category / contract_type
        # emp_category = str(get('Employeement Type') or '').strip()
        # if emp_category:
        #     employee_vals['employee_category'] = emp_category

        # Create employee
        try:
            employee = env['hr.employee'].sudo().create(employee_vals)
            # Pro trick: explicitly write phone fields after creation to bypass any compute/inverse race conditions that wipe them
            phone_vals = {}
            if employee_vals.get('work_phone'):
                phone_vals['work_phone'] = employee_vals.get('work_phone')
            if employee_vals.get('mobile_phone'):
                phone_vals['mobile_phone'] = employee_vals.get('mobile_phone')
            if employee_vals.get('private_phone'):
                phone_vals['private_phone'] = employee_vals.get('private_phone')
            if phone_vals:
                employee.sudo().write(phone_vals)
        except Exception as exc:
            return {'error': f'Failed to create employee: {exc}'}

        # --------------- Create Bank Accounts ---------------
        bank_accounts_str = _parse_str(get('Bank Accounts'))
        if bank_accounts_str:
            bank_acc_list = [b.strip() for b in bank_accounts_str.split(',') if b.strip()]
            for acc_number in bank_acc_list:
                # Find partner for the bank account (work_contact_id or company.partner_id)
                partner_id = getattr(employee, 'work_contact_id', False) and employee.work_contact_id.id or employee.company_id.partner_id.id

                # Look for existing bank account
                bank_acc = env['res.partner.bank'].sudo().search([('acc_number', '=', acc_number)], limit=1)
                if not bank_acc:
                    bank_vals = {
                        'acc_number': acc_number,
                        'company_id': employee.company_id.id,
                    }
                    if partner_id:
                        bank_vals['partner_id'] = partner_id
                    bank_acc = env['res.partner.bank'].sudo().create(bank_vals)
                employee.sudo().write({'bank_account_ids': [(4, bank_acc.id)]})

        # --------------- Create res.users ---------------
        password = _random_password()
        try:
            # Gather groups: 'Internal User' (base.group_user) + 'Normal Employees'
            groups_list = [env.ref('base.group_user').id]
            normal_emp_group = env['res.groups'].sudo().search([('name', '=', 'Normal Employees')], limit=1)
            if normal_emp_group:
                groups_list.append(normal_emp_group.id)

            user_vals = {
                'name': emp_name,
                'login': work_email,
                'email': work_email,
                'company_id': company.id,
                'company_ids': [(4, company.id)],
                'group_ids': [(6, 0, groups_list)],
            }
            user = env['res.users'].sudo().create(user_vals)
            # Set password without sending reset email
            user.sudo()._set_encrypted_password(
                user.id,
                env['res.users']._crypt_context().hash(password),
            )
            employee.sudo().write({'user_id': user.id})
        except Exception as exc:
            return {'error': f'Failed to create user for employee: {exc}'}

        return {
            'employee_name': emp_name,
            'login_email': work_email,
            'password': password,
            'company': company_name,
            'department': dept_name,
            'designation': job_name,
        }

    # -----------------------------------------------------------------------
    # Private – Helpers
    # -----------------------------------------------------------------------

    def _resolve_employee_ref(self, ref):
        """
        Resolve a supervisor/manager reference.
        `ref` can be an Employee ID (employee_id_no) or Work Email.
        Returns employee.id or None.
        """
        if not ref:
            return None
        # Try by employee_id_no first
        emp = self.env['hr.employee'].sudo().search(
            [('employee_id_no', '=', ref)], limit=1
        )
        if emp:
            return emp.id
        # Try by work_email
        emp = self.env['hr.employee'].sudo().search(
            [('work_email', '=', ref)], limit=1
        )
        return emp.id if emp else None

    # -----------------------------------------------------------------------
    # Private – Excel builders
    # -----------------------------------------------------------------------

    def _build_output_excel(self, results):
        """Build the password output Excel file. Returns (b64_content, filename)."""
        if not results:
            return False, False
            
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_('openpyxl is required. Install it with: pip install openpyxl'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Imported Employees'

        headers = [
            'Employee Name', 'Login Email', 'Password',
            'Company', 'Department', 'Designation',
        ]

        # Header row styling
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 28

        for row_idx, r in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=r.get('employee_name', ''))
            ws.cell(row=row_idx, column=2, value=r.get('login_email', ''))
            ws.cell(row=row_idx, column=3, value=r.get('password', ''))
            ws.cell(row=row_idx, column=4, value=r.get('company', ''))
            ws.cell(row=row_idx, column=5, value=r.get('department', ''))
            ws.cell(row=row_idx, column=6, value=r.get('designation', ''))

        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()
        fname = f'employee_import_output_{fields.Date.today()}.xlsx'
        return b64, fname

    def _build_error_excel(self, failed_rows):
        """Build the error report Excel. Returns (b64_content, filename)."""
        if not failed_rows:
            return False, False

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_('openpyxl is required.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Errors'

        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx, h in enumerate(['Row #', 'Error Details'], start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 80

        for row_idx, fr in enumerate(failed_rows, start=2):
            ws.cell(row=row_idx, column=1, value=fr.get('row_no', ''))
            ws.cell(row=row_idx, column=2, value=fr.get('error', ''))

        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()
        fname = f'employee_import_errors_{fields.Date.today()}.xlsx'
        return b64, fname
